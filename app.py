from flask import Flask, request, send_file, jsonify, render_template_string
from werkzeug.utils import secure_filename
from automacao_esp import processar_nota
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
PLANILHA_PATH = "dados.xlsx"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Criação da planilha se não existir
if not os.path.exists(PLANILHA_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nota Fiscal", "Série", "Matrícula", "CPF/CNPJ", "Nome Destinatário"])
    wb.save(PLANILHA_PATH)

@app.route("/")
def index():
    return render_template_string("""
        <h2>Robo Espaider - Upload de Notas Fiscais</h2>
        <form action="/processar" method="post" enctype="multipart/form-data">
            Usuário: <input type="text" name="usuario"><br><br>
            Senha: <input type="password" name="senha"><br><br>
            Selecione os arquivos PDF: <input type="file" name="arquivo" multiple><br><br>
            <input type="submit" value="Enviar">
        </form>
        <br>
        <a href="/baixar-planilha">Baixar Planilha Excel</a>
    """)

@app.route("/processar", methods=["POST"])
def processar():
    usuario = request.form.get("usuario")
    senha = request.form.get("senha")

    files = request.files.getlist("arquivo")

    wb = load_workbook(PLANILHA_PATH)
    ws = wb.active

    for file in files:
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        resultado = processar_nota(filepath, usuario, senha)

        ws.append([
            resultado["numero_nota"],
            resultado["serie"],
            resultado["matricula"],
            resultado["cpf_cnpj"],
            resultado["nome"]
        ])

    wb.save(PLANILHA_PATH)

    return "Processamento concluído com sucesso! <a href='/'>Voltar</a>"

@app.route("/baixar-planilha", methods=["GET"])
def baixar():
    return send_file(PLANILHA_PATH, as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
